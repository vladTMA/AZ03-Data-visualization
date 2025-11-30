import pytest
from myproject import saver
import openpyxl

@pytest.mark.unit
def test_save_xlsx(tmp_path):
    data = [{"name": "Пример", "price": 1000, "currency": "RUB", "url": "https://example.com"}]
    saver.OUTPUT_DIR = tmp_path
    saver.save_xlsx(data, filename="test.xlsx")

    file_path = tmp_path / "test.xlsx"
    assert file_path.exists()

    # Проверяем содержимое XLSX
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    # Используются русские названия заголовков
    assert "Название" in headers and "Ссылка" in headers and "Цена" in headers and "Валюта" in headers
    # Находим индексы колонок
    name_idx = headers.index("Название") + 1
    price_idx = headers.index("Цена") + 1
    currency_idx = headers.index("Валюта") + 1
    url_idx = headers.index("Ссылка") + 1
    assert sheet.cell(row=2, column=name_idx).value == "Пример"
    assert sheet.cell(row=2, column=price_idx).value == 1000
    assert sheet.cell(row=2, column=currency_idx).value == "RUB"
    assert sheet.cell(row=2, column=url_idx).value == "https://example.com"
