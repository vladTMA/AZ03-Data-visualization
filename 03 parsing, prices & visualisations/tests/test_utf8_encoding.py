import pytest
import openpyxl
from myproject import saver


@pytest.mark.encoding
def test_utf8_xlsx_encoding(tmp_path):
    data = [{"name": "Тестовая строка", "price": 1000, "currency": "RUB", "url": "https://пример.ру"}]
    saver.OUTPUT_DIR = tmp_path
    saver.save_xlsx(data, filename="utf8.xlsx")

    xlsx_file = tmp_path / "utf8.xlsx"
    wb = openpyxl.load_workbook(xlsx_file)
    sheet = wb.active

    # Проверяем, что кириллица сохранилась
    # Структура колонок: name (1), price (2), currency (3), url (4)
    headers = [cell.value for cell in sheet[1]]
    name_idx = headers.index("Название") + 1
    url_idx = headers.index("Ссылка") + 1
    
    assert sheet.cell(row=2, column=name_idx).value == "Тестовая строка"
    assert sheet.cell(row=2, column=url_idx).value == "https://пример.ру"


